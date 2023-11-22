#author : PRiyanka Yadhav Suresh
#data source : Any excel sheet of your choice 
#Topic : Reading data in excel
import openpyxl 

#loading the workbook
data_workbook = openpyxl.load_workbook('news.xlsx')
worksheet = data_workbook.active
worksheet = data_workbook['news']
print('Total number of rows: '+str(worksheet.max_row)+'. And total number of columns: '+str(worksheet.max_column))

#Reading Data from one cell
print('The value in cell A1 is: '+worksheet['A1'].value)

#Reading from multiple cells
values = [worksheet.cell(row=1,column=i).value for i in range(1,worksheet.max_column+1)]
print(values)
data=[worksheet.cell(row=i,column=2).value for i in range(2,12)]
print(data)

#to print first 10 rows
data_list = list()

for value in worksheet.iter_rows(min_row=1, max_row=11, min_col=1, max_col=6, values_only=True):
    data_list.append(value)
    
for ele1,ele2,ele3,ele4,ele5,ele6 in data_list:
    if ele1 is not None and ele2 is not None and ele3 is not None and ele4 is not None and ele5 is not None and ele6 is not None:
     print("{:<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1, ele2, ele3, ele4, ele5, ele6))
else:
    print("One or more variables is None and cannot be formatted.")

   